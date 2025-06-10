import re
import os
import sys
import openpyxl
from openpyxl.styles import Font
import argparse 
import datetime 

# --- Helper Functions ---
def extract_ip_addresses(lines):
    # Pattern for valid IPv4 addresses (e.g., 192.168.1.1)
    ip_pattern = re.compile(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})') 
    return [match.group(1) for line in lines for match in ip_pattern.finditer(line)]

def determine_interface_mode(lines):
    # This function is primarily used for IOS-XE and the initial basic mode/vlan detection.
    # For NX-OS, _finalize_nxos_interface_data now performs more comprehensive extraction.
    mode = "Unknown"
    vlan = "None"
    for line in lines:
        if "switchport mode access" in line:
            mode = "Access"
        elif "switchport mode trunk" in line:
            mode = "Trunk"
        elif "no switchport" in line: # For Layer 3 interfaces (routed)
            mode = "Routed"
        if "access vlan" in line:
            vlan_match = re.search(r'access vlan (\d+)', line)
            if vlan_match:
                vlan = vlan_match.group(1)
        elif "trunk allowed vlan" in line:
            vlan_match = re.search(r'trunk allowed vlan ([\d,\-]+)', line)
            if vlan_match:
                vlan = vlan_match.group(1)
    return mode, vlan
# --- End Helper Functions ---


# --- OS Specific Parsers ---

def _parse_ios_xe_config(config_lines):
    parsed_data = {
        "os_type": "IOS-XE",
        "hostname": None,
        "interfaces": [],
        "routing": [], 
        "static_routes": [], 
        "vlans": [],
        "ntp": [],
        "snmp": [],
        "syslog": [],
        "radius": [],
    }

    current_interface = None
    current_radius_context_name = None 

    i = 0
    while i < len(config_lines):
        line = config_lines[i].strip() 

        # --- Context Clearing Logic ---
        is_new_top_level_command = any(line.startswith(cmd) for cmd in [
            "hostname", "interface", "router", "vlan",
            "ip route", "ntp server", "snmp-server host", "logging host",
            "radius server", "radius-server host", "aaa group server radius" 
        ])

        if current_interface and (line.startswith("!") or (is_new_top_level_command and not line.startswith("interface"))):
            parsed_data["interfaces"].append(current_interface)
            current_interface = None 

        if current_radius_context_name and (line.startswith("!") or is_new_top_level_command) \
           and not any(line.startswith(cmd) for cmd in ["radius server", "radius-server host", "aaa group server radius"]):
            current_radius_context_name = None
        # --- End Context Clearing Logic ---
        
        if line.startswith("hostname"):
            parsed_data["hostname"] = line.split()[1]

        elif line.startswith("interface"):
            if current_interface: 
                parsed_data["interfaces"].append(current_interface)
            current_interface = {"name": line.split()[1], "lines": []}

        elif current_interface:
            current_interface["lines"].append(line)

        elif line.startswith("router"):
            parsed_data["routing"].extend(extract_ip_addresses([line]))

        # --- IOS-XE Static Routes ---
        elif line.startswith("ip route"):
            route_match = re.search(
                r'^ip route(?: vrf (\S+))?\s+' 
                r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})' 
                r'(?:\/(\d{1,2})|\s+(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}))\s+' 
                r'(\S+)' 
                r'(?: (\d+))?$', 
                line
            )
            if route_match:
                vrf = route_match.group(1) if route_match.group(1) else "default"
                destination = route_match.group(2)
                
                if route_match.group(3): 
                    mask = f"/{route_match.group(3)}"
                elif route_match.group(4): 
                    mask = route_match.group(4)
                else: 
                    mask = "N/A"

                next_hop_or_interface = route_match.group(5)
                metric = route_match.group(6) if route_match.group(6) else "N/A"

                route_data = {
                    "vrf": vrf,
                    "destination": destination,
                    "mask": mask, 
                    "next_hop_or_interface": next_hop_or_interface,
                    "metric": metric,
                    "type": "static_ip_route" 
                }
                parsed_data["static_routes"].append(route_data)

        elif line.startswith("vlan"):
            parsed_data["vlans"].append(line) # IOS-XE VLAN parsing remains simple, no nested 'name' context needed here.

        elif line.startswith("ntp server"): 
            extracted_ips = extract_ip_addresses([line]) 
            parsed_data["ntp"].extend(extracted_ips)

        elif line.startswith("snmp-server host"):
            extracted_ips = extract_ip_addresses([line]) 
            parsed_data["snmp"].extend(extracted_ips)

        elif line.startswith("logging host"):
            extracted_ips = extract_ip_addresses([line]) 
            parsed_data["syslog"].extend(extracted_ips)

        # RADIUS Parsing Logic
        elif line.startswith("radius server"):
            name_match = re.search(r'radius server (\S+)', line)
            if name_match:
                current_radius_context_name = name_match.group(1)
            else:
                current_radius_context_name = None 
        
        elif line.startswith("address ipv4") and current_radius_context_name: 
            ip_match = re.search(r'address ipv4 (\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', line)
            if ip_match:
                parsed_data["radius"].append((current_radius_context_name, ip_match.group(1)))
                current_radius_context_name = None 
        
        elif line.startswith("radius-server host"):
            ip_match = re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', line)
            name_match = re.search(r'radius-server host (\S+)', line) 
            if ip_match:
                server_name = name_match.group(1) if name_match else "Unknown_Old_Format"
                parsed_data["radius"].append((server_name, ip_match.group(1)))
            current_radius_context_name = None 

        i += 1
    
    if current_interface:
        parsed_data["interfaces"].append(current_interface)

    return parsed_data

# --- NX-OS Specific Interface/VLAN Finalizers ---
def _finalize_nxos_interface_data(iface_dict, parsed_data):
    """
    Helper to process an NX-OS interface's collected lines before adding it to parsed_data.
    Extracts mode, vlan, description, IP info, channel-group, and VPC ID if applicable.
    """
    # Extract IP info (already done in main parse loop for IOS-XE, but centralizing for NX-OS)
    ip_lines = extract_ip_addresses(iface_dict['lines'])
    iface_dict["ip_info"] = ", ".join(ip_lines) if ip_lines else "No IP"

    # Extract mode and vlan info
    mode, vlan = determine_interface_mode(iface_dict["lines"]) # Re-using for base mode/vlan string
    iface_dict["mode"] = mode
    # iface_dict["vlan"] will be set by specific logic below if needed.
    
    # Extract description
    description_match = next((re.search(r'description (.+)', line) for line in iface_dict['lines'] if re.match(r'description ', line)), None)
    iface_dict["description"] = description_match.group(1) if description_match else "No Description"

    # --- NEW: Extract Channel-group info ---
    channel_group_match = None
    for sub_line in iface_dict["lines"]:
        stripped_sub_line = sub_line.strip()
        if stripped_sub_line.startswith("channel-group "):
            # Example: channel-group 1 mode active
            cg_match = re.search(r'channel-group (\d+) mode (\S+)', stripped_sub_line)
            if cg_match:
                iface_dict["channel_group"] = {"id": cg_match.group(1), "mode": cg_match.group(2)}
                break # Found it, no need to search further
    # --- END NEW ---

    # Extract VPC ID for Port-Channels (NX-OS specific)
    if iface_dict["name"].lower().startswith("port-channel"):
        for sub_line in iface_dict["lines"]:
            stripped_sub_line = sub_line.strip()
            if stripped_sub_line.startswith("vpc "):
                vpc_id_match = re.search(r'vpc (\d+)', stripped_sub_line)
                if vpc_id_match:
                    iface_dict["vpc_id"] = vpc_id_match.group(1)
                    break # Found VPC ID, no need to search further

    # --- NEW: Extract all 'switchport trunk allowed vlan' lines ---
    allowed_vlans_config_list = []
    for sub_line in iface_dict["lines"]:
        stripped_sub_line = sub_line.strip()
        # Captures 'switchport trunk allowed vlan 10,20', 'switchport trunk allowed vlan add 30', etc.
        if stripped_sub_line.startswith("switchport trunk allowed vlan"):
            allowed_vlans_config_list.append(stripped_sub_line)
    if allowed_vlans_config_list:
        iface_dict["allowed_vlans_config"] = allowed_vlans_config_list
        # If specific allowed_vlans_config exist, override the simple 'vlan' string
        iface_dict["vlan"] = f"Configured: {'; '.join(allowed_vlans_config_list)}"
    else:
        # If no specific allowed_vlans_config, use the basic vlan string from determine_interface_mode
        iface_dict["vlan"] = vlan 
    # --- END NEW ---

    # Add the finalized interface data to the main parsed_data
    parsed_data["interfaces"].append(iface_dict)


def _parse_nx_os_config(config_lines):
    parsed_data = {
        "os_type": "NX-OS",
        "hostname": None,
        "interfaces": [], # Interfaces will be built by _finalize_nxos_interface_data
        "routing": [], 
        "static_routes": [], 
        "vlans": [], # Will store dictionaries with 'id' and 'name'
        "ntp": [],
        "snmp": [],
        "syslog": [],
        "radius": [],
        "features": [], 
        "vpc": {}, # Stores global VPC domain config, keyed by domain ID
    }

    current_interface = None
    current_radius_context_name = None 
    current_vpc_domain = None 
    current_vlan_data = None # New: to track current VLAN block for its name

    i = 0
    while i < len(config_lines):
        line = config_lines[i].strip()

        # --- Context Clearing Logic ---
        is_new_top_level_command = any(line.startswith(cmd) for cmd in [
            "hostname", "interface", "vlan", "feature", "vpc domain",
            "ip route", "ntp server", "snmp-server host", "logging server",
            "radius server", "radius-server host", "aaa group server radius" 
        ])

        # 1. Handle closing of `current_interface` context
        if current_interface and (line.startswith("!") or (is_new_top_level_command and not line.startswith("interface"))):
            _finalize_nxos_interface_data(current_interface, parsed_data) # Finalize before clearing
            current_interface = None 

        # 2. Handle closing of `current_radius_context_name` context
        if current_radius_context_name and (line.startswith("!") or is_new_top_level_command) \
           and not any(line.startswith(cmd) for cmd in ["radius server", "radius-server host", "aaa group server radius"]):
            current_radius_context_name = None
        
        # 3. Handle closing of `current_vpc_domain` context
        if current_vpc_domain and (line.startswith("!") or (is_new_top_level_command and not line.startswith("vpc domain"))):
            current_vpc_domain = None

        # 4. Handle closing of `current_vlan_data` context (for 'vlan <id>' blocks with nested 'name')
        if current_vlan_data and (line.startswith("!") or (is_new_top_level_command and not line.startswith("vlan "))): # Note: "vlan " with space to avoid matching partials like "vlan-group"
            parsed_data["vlans"].append(current_vlan_data) # Add finalized VLAN data
            current_vlan_data = None
        # --- End Context Clearing Logic ---

        if line.startswith("hostname"):
            parsed_data["hostname"] = line.split()[1]

        elif line.startswith("interface"):
            if current_interface: # This should ideally be handled by the context clearing above
                _finalize_nxos_interface_data(current_interface, parsed_data) 
            current_interface = {"name": line.split()[1], "lines": []}
            
        elif current_interface:
            current_interface["lines"].append(line)
        
        # --- NEW: VLAN Parsing for NX-OS (as context, handles 'name' line) ---
        elif line.startswith("vlan "): 
            if current_vlan_data: # If previous VLAN block was not closed by context clearing (fallback)
                parsed_data["vlans"].append(current_vlan_data)
            
            vlan_id_match = re.search(r'vlan ([\d,-]+)', line) # Capture ID (and ranges like 1-5,10)
            vlan_id = vlan_id_match.group(1) if vlan_id_match else "Unknown_VLAN"
            current_vlan_data = {"id": vlan_id, "name": "N/A"} # Start new VLAN context
        
        elif current_vlan_data: # If inside a VLAN block
            if line.startswith("name "):
                current_vlan_data["name"] = line.split(" ", 1)[1]
            # Add other VLAN sub-commands here if needed (e.g., state active/suspend)
        # --- END NEW: VLAN Parsing ---

        elif line.startswith("feature "):
            parsed_data["features"].append(line.split(" ", 1)[1]) 
        
        elif line.startswith("ntp server"):
            extracted_ips = extract_ip_addresses([line]) 
            if extracted_ips:
                parsed_data["ntp"].extend(extracted_ips)

        elif line.startswith("snmp-server host"):
            extracted_ips = extract_ip_addresses([line]) 
            if extracted_ips:
                parsed_data["snmp"].extend(extracted_ips)
        
        elif line.startswith("logging server"): 
            extracted_ips = extract_ip_addresses([line]) 
            if extracted_ips:
                parsed_data["syslog"].extend(extracted_ips)

        # --- NX-OS Static Routes ---
        elif line.startswith("ip route"):
            route_match = re.search(
                r'^ip route(?: vrf (\S+))?\s+' 
                r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\/\d{1,2})\s+' 
                r'(\S+)' 
                r'(?: pref (\d+))?$', 
                line
            )
            if route_match:
                vrf = route_match.group(1) if route_match.group(1) else "default"
                destination_cidr = route_match.group(2)
                next_hop_or_interface = route_match.group(3)
                preference = route_match.group(4) if route_match.group(4) else "N/A"

                route_data = {
                    "vrf": vrf,
                    "destination_cidr": destination_cidr,
                    "next_hop_or_interface": next_hop_or_interface,
                    "preference": preference,
                    "type": "static_ip_route"
                }
                parsed_data["static_routes"].append(route_data)

        # --- VPC Domain Parsing for NX-OS ---
        elif line.startswith("vpc domain "):
            domain_id_match = re.search(r'vpc domain (\d+)', line)
            domain_id = domain_id_match.group(1) if domain_id_match else "Unknown_Domain"
            current_vpc_domain = {"domain_id": domain_id, "settings": {}}
            parsed_data["vpc"][domain_id] = current_vpc_domain 
        
        elif current_vpc_domain: # If we are inside a VPC domain block, parse its sub-commands
            if line.startswith("role "):
                current_vpc_domain["settings"]["role"] = line.split(" ", 1)[1]
            elif line.startswith("system-priority "):
                current_vpc_domain["settings"]["system-priority"] = line.split(" ", 1)[1]
            elif line.startswith("peer-keepalive destination "):
                peer_match = re.search(r'peer-keepalive destination (\S+) source (\S+)', line)
                if peer_match:
                    current_vpc_domain["settings"]["peer-keepalive_destination"] = peer_match.group(1)
                    current_vpc_domain["settings"]["peer-keepalive_source"] = peer_match.group(2)
            elif line.startswith("peer-gateway"):
                current_vpc_domain["settings"]["peer-gateway"] = "enabled"
            elif line.startswith("auto-recovery"):
                current_vpc_domain["settings"]["auto-recovery"] = "enabled"
            elif line.startswith("delay restore "):
                current_vpc_domain["settings"]["delay-restore"] = line.split(" ", 1)[1]
            elif line.startswith("peer-switch"):
                current_vpc_domain["settings"]["peer-switch"] = "enabled"
        # --- END VPC Domain Parsing ---

        # RADIUS for NX-OS
        elif line.startswith("aaa group server radius"):
            group_match = re.search(r'aaa group server radius (\S+)', line)
            if group_match:
                current_radius_context_name = group_match.group(1)
        
        elif line.startswith("server ") and current_radius_context_name: 
            ip_match = re.search(r'server (\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', line)
            if ip_match:
                parsed_data["radius"].append((current_radius_context_name, ip_match.group(1)))

        elif line.startswith("radius server"): 
            name_match = re.search(r'radius server (\S+)', line)
            if name_match:
                current_radius_context_name = name_match.group(1)
        
        elif line.startswith("address ipv4") and current_radius_context_name: 
            ip_match = re.search(r'address ipv4 (\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', line)
            if ip_match:
                parsed_data["radius"].append((current_radius_context_name, ip_match.group(1)))
                current_radius_context_name = None 
        
        elif line.startswith("radius-server host"): 
            ip_match = re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', line)
            name_match = re.search(r'radius-server host (\S+)', line) 
            if ip_match:
                server_name = name_match.group(1) if name_match else "Unknown_Old_Format"
                parsed_data["radius"].append((server_name, ip_match.group(1)))
            current_radius_context_name = None 

        i += 1
    
    # Finalize any lingering contexts after loop ends
    if current_interface: 
        _finalize_nxos_interface_data(current_interface, parsed_data)
    if current_vlan_data: # NEW: Finalize last VLAN data if any
        parsed_data["vlans"].append(current_vlan_data)

    return parsed_data

# --- OS Detection Logic ---
def detect_os_type(config_lines):
    """
    Analyzes the config lines to determine if it's IOS-XE or NX-OS.
    """
    lines_to_check = config_lines[:100] # Check the first 100 lines for indicators

    for line in lines_to_check: 
        stripped_line = line.strip()
        # Highly specific NX-OS indicators
        if stripped_line.startswith("feature "):
            return "NX-OS"
        if stripped_line.startswith("interface Ethernet") and not "interface Ethernet-Controller" in stripped_line:
            return "NX-OS"
        if stripped_line.startswith("vpc domain"): 
            return "NX-OS"
        if stripped_line.startswith("vrf context"):
            return "NX-OS"
        if stripped_line.startswith("install all"):
            return "NX-OS"
        if stripped_line.startswith("install feature-set"):
            return "NX-OS"
        if stripped_line.startswith("redundancy role"):
            return "NX-OS"
        if stripped_line.startswith("hardware profile"):
            return "NX-OS"
        # Check for 'NX-OS' explicitly in the line (e.g., in 'show version' output or comments)
        if "NX-OS" in stripped_line.upper(): 
            return "NX-OS"
    
    # If no strong NX-OS indicators found after checking sufficient lines, assume IOS-XE
    return "IOS-XE"


# --- Main Parsing Orchestrator ---
def parse_config(config_lines):
    """
    Detects OS type and calls the appropriate parsing function.
    """
    os_type = detect_os_type(config_lines)
    print(f"Detected OS Type: {os_type}") # Informative for the user
    if os_type == "IOS-XE":
        return _parse_ios_xe_config(config_lines)
    elif os_type == "NX-OS":
        return _parse_nx_os_config(config_lines)
    else:
        raise ValueError("Could not determine OS type from configuration.")


# --- Export to Excel Function ---
def export_to_excel(parsed_data, output_file, collection_date_obj): 
    """
    Exports the parsed configuration data for a single device to its own Excel worksheet.
    It will create a new sheet named after the device's hostname, replacing it if it already exists.
    """
    hostname = parsed_data["hostname"]
    if not hostname:
        print("Warning: Hostname not found in configuration. Using 'Unknown_Device' for sheet name.")
        hostname = "Unknown_Device"

    wb = None
    if os.path.exists(output_file) and os.path.getsize(output_file) > 0:
        try:
            wb = openpyxl.load_workbook(output_file)
        except openpyxl.utils.exceptions.InvalidFileException:
            print(f"Warning: '{output_file}' exists but is not a valid Excel file. Creating a new one.")
            wb = openpyxl.Workbook()
    
    if wb is None: 
        wb = openpyxl.Workbook()

    # --- Sheet Management ---
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1 and not wb['Sheet'].max_row > 0:
        wb.remove(wb['Sheet'])
    if hostname in wb.sheetnames:
        wb.remove(wb[hostname])
    
    ws = wb.create_sheet(title=hostname)
    # --- End Sheet Management ---

    # --- Set Headers ---
    ws.cell(row=1, column=1).value = "Descriptor"
    ws.cell(row=1, column=2).value = "Configuration"
    for col in range(1, 3): 
        ws.cell(row=1, column=col).font = Font(bold=True)
    # --- End Set Headers ---

    current_sheet_row_num = 2 

    def write_row_to_sheet(descriptor, value):
        nonlocal current_sheet_row_num
        ws.cell(row=current_sheet_row_num, column=1).value = descriptor
        ws.cell(row=current_sheet_row_num, column=2).value = value
        current_sheet_row_num += 1

    # --- Write Hostname & OS Type & Collection Date ---
    write_row_to_sheet("Hostname", parsed_data["hostname"] or "Not Found")
    write_row_to_sheet("OS Type", parsed_data["os_type"] or "Unknown")
    write_row_to_sheet("Collection Date", collection_date_obj.strftime("%Y-%m-%d")) 

    # --- Write Features (NX-OS Specific) ---
    if parsed_data.get("features") and parsed_data["os_type"] == "NX-OS":
        ws.cell(row=current_sheet_row_num, column=1).value = "--- FEATURES ---"
        ws.cell(row=current_sheet_row_num, column=1).font = Font(bold=True)
        current_sheet_row_num += 1
        for i, feature in enumerate(parsed_data["features"], start=1):
            write_row_to_sheet(f"Feature #{i}", feature)

    # --- Write VPC Domain Configuration (NX-OS Specific) ---
    if parsed_data.get("vpc") and parsed_data["os_type"] == "NX-OS":
        ws.cell(row=current_sheet_row_num, column=1).value = "--- VPC DOMAINS ---"
        ws.cell(row=current_sheet_row_num, column=1).font = Font(bold=True)
        current_sheet_row_num += 1
        for domain_id, domain_data in parsed_data["vpc"].items():
            write_row_to_sheet(f"VPC Domain {domain_id}", "") # Empty value for header row
            for setting_name, setting_value in sorted(domain_data["settings"].items()):
                # Format boolean/enabled values nicely
                display_value = str(setting_value).replace("_", " ").title() if isinstance(setting_value, bool) else str(setting_value)
                write_row_to_sheet(f"  {setting_name.replace('-', ' ').title()}", display_value)
            current_sheet_row_num += 1 # Add extra space after each domain for readability

    # --- Write Interfaces ---
    ws.cell(row=current_sheet_row_num, column=1).value = "--- INTERFACES ---"
    ws.cell(row=current_sheet_row_num, column=1).font = Font(bold=True)
    current_sheet_row_num += 1

    # Iterate through finalized interface data
    for iface in parsed_data["interfaces"]:
        # All necessary details are now directly on the iface dictionary for NX-OS
        ip_info = iface.get("ip_info", "No IP") 
        mode = iface.get("mode", "Unknown")
        vlan = iface.get("vlan", "None") # Now contains detailed info for NX-OS
        description = iface.get("description", "No Description")
        vpc_info_on_interface = f" | VPC: {iface['vpc_id']}" if iface.get("vpc_id") else "" 
        channel_group_info = f" | CG: {iface['channel_group']['id']} ({iface['channel_group']['mode']})" if iface.get("channel_group") else "" # NEW: Channel-group info

        write_row_to_sheet(f"Interface {iface['name']}", f"IP: {ip_info} | Mode: {mode} | VLAN: {vlan} | Desc: {description}{vpc_info_on_interface}{channel_group_info}") # Combined string

    # --- Write Routing Information (Generic, but mostly applies to OSPF/EIGRP network statements) ---
    if parsed_data["routing"]:
        ws.cell(row=current_sheet_row_num, column=1).value = "--- ROUTING IPs ---"
        ws.cell(row=current_sheet_row_num, column=1).font = Font(bold=True)
        current_sheet_row_num += 1
        for i, ip in enumerate(parsed_data["routing"], start=1):
            write_row_to_sheet(f"Routing IP #{i}", ip)

    # --- Write Static Routes ---
    if parsed_data["static_routes"]:
        ws.cell(row=current_sheet_row_num, column=1).value = "--- STATIC ROUTES ---"
        ws.cell(row=current_sheet_row_num, column=1).font = Font(bold=True)
        current_sheet_row_num += 1
        for i, route in enumerate(parsed_data["static_routes"], start=1):
            if parsed_data["os_type"] == "IOS-XE":
                route_str = (f"VRF: {route['vrf']} | Dest: {route['destination']} {route['mask']} | "
                             f"Next-Hop/Int: {route['next_hop_or_interface']} | Metric: {route['metric']}")
            elif parsed_data["os_type"] == "NX-OS":
                route_str = (f"VRF: {route['vrf']} | Dest: {route['destination_cidr']} | "
                             f"Next-Hop/Int: {route['next_hop_or_interface']} | Preference: {route['preference']}")
            else: 
                route_str = str(route) 
            write_row_to_sheet(f"Static Route #{i}", route_str)

    # --- Write VLANs ---
    if parsed_data["vlans"]:
        ws.cell(row=current_sheet_row_num, column=1).value = "--- VLANs ---"
        ws.cell(row=current_sheet_row_num, column=1).font = Font(bold=True)
        current_sheet_row_num += 1
        for i, vlan_data in enumerate(parsed_data["vlans"], start=1): # Iterating through dicts now
            write_row_to_sheet(f"VLAN {vlan_data['id']}", f"Name: {vlan_data['name']}") # Access name from dict

    # --- Write NTP Servers ---
    if parsed_data["ntp"]:
        ws.cell(row=current_sheet_row_num, column=1).value = "--- NTP SERVERS ---"
        ws.cell(row=current_sheet_row_num, column=1).font = Font(bold=True)
        current_sheet_row_num += 1
        for i, ip in enumerate(parsed_data["ntp"], start=1):
            write_row_to_sheet(f"NTP Server #{i}", ip)

    # --- Write SNMP Hosts ---
    if parsed_data["snmp"]:
        ws.cell(row=current_sheet_row_num, column=1).value = "--- SNMP HOSTS ---"
        ws.cell(row=current_sheet_row_num, column=1).font = Font(bold=True)
        current_sheet_row_num += 1
        for i, ip in enumerate(parsed_data["snmp"], start=1):
            write_row_to_sheet(f"SNMP Host #{i}", ip)

    # --- Write Syslog Servers ---
    if parsed_data["syslog"]:
        ws.cell(row=current_sheet_row_num, column=1).value = "--- SYSLOG SERVERS ---"
        ws.cell(row=current_sheet_row_num, column=1).font = Font(bold=True)
        current_sheet_row_num += 1
        for i, ip in enumerate(parsed_data["syslog"], start=1):
            write_row_to_sheet(f"Syslog Server #{i}", ip)

    # --- Write RADIUS Hosts ---
    if parsed_data["radius"]:
        ws.cell(row=current_sheet_row_num, column=1).value = "--- RADIUS HOSTS ---"
        ws.cell(row=current_sheet_row_num, column=1).font = Font(bold=True)
        current_sheet_row_num += 1
        for i, (name, ip) in enumerate(parsed_data["radius"], start=1):
            write_row_to_sheet(f"RADIUS Host #{i}", f"{name} ({ip})")
    
    wb.save(output_file)


def main():
    # Setup argument parser
    parser = argparse.ArgumentParser(description="Parse Cisco IOS-XE/NX-OS configurations and export to Excel.")
    parser.add_argument("path", help="Path to a single config file (.txt) or a directory containing config files.")
    parser.add_argument("-r", "--recursive", action="store_true", 
                        help="Read all .txt files from the specified directory. 'path' must be a directory.")
    parser.add_argument("-o", "--output", default="network_config_export.xlsx",
                        help="Specify the name of the Excel output file (default: network_config_export.xlsx).")
    args = parser.parse_args()

    # Get the current date for the export
    current_date = datetime.date.today()

    files_to_process = []

    if args.recursive:
        if not os.path.isdir(args.path):
            print(f"Error: Directory '{args.path}' not found or is not a directory for recursive mode.")
            sys.exit(1)
        
        for filename in os.listdir(args.path):
            if filename.endswith(".txt"):
                files_to_process.append(os.path.join(args.path, filename))
        
        if not files_to_process:
            print(f"No .txt files found in directory '{args.path}'.")
            sys.exit(0)

    else: # Single file mode
        if not os.path.isfile(args.path):
            print(f"Error: File '{args.path}' not found.")
            sys.exit(1)
        files_to_process.append(args.path)

    output_file = args.output

    print(f"Starting configuration export to '{output_file}'...")

    for file_path in files_to_process:
        print(f"\nProcessing file: '{file_path}'...")
        try:
            with open(file_path, 'r') as f:
                config_lines = f.readlines()
        except Exception as e:
            print(f"Error reading file '{file_path}': {e}. Skipping.")
            continue

        try:
            parsed = parse_config(config_lines)
            # Pass the current_date object to the export function
            export_to_excel(parsed, output_file, current_date)
            print(f"Successfully exported configuration for '{parsed['hostname'] or 'Unknown_Device'}' ({parsed['os_type']}).")
        except Exception as e:
            print(f"Error parsing or exporting config from '{file_path}': {e}. Skipping.")
            # For debugging errors during parsing/export, you might want to uncomment traceback.print_exc()
            # import traceback
            # traceback.print_exc()
            continue

    print("\nAll specified configuration files processed. Check the Excel file for results.")


if __name__ == "__main__":
    main()
