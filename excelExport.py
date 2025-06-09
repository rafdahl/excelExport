import re
import os
import sys
import openpyxl
from openpyxl.styles import Font

def extract_ip_addresses(lines):
    """
    Extracts IPv4 addresses from a list of lines using a regular expression.
    """
    ip_pattern = re.compile(r'(\d+\.\d+\.\d+\.\d+)')
    return [match.group(1) for line in lines for match in ip_pattern.finditer(line)]

def determine_interface_mode(lines):
    """
    Determines the switchport mode (Access, Trunk, Routed) and associated VLAN(s)
    from a list of interface configuration lines.
    """
    mode = "Unknown"
    vlan = "None"
    for line in lines:
        if "switchport mode access" in line:
            mode = "Access"
        elif "switchport mode trunk" in line:
            mode = "Trunk"
        elif "no switchport" in line:
            mode = "Routed"
        if "access vlan" in line:
            vlan_match = re.search(r'access vlan (\d+)', line)
            if vlan_match:
                vlan = vlan_match.group(1)
        elif "trunk allowed vlan" in line:
            vlan_match = re.search(r'trunk allowed vlan ([\d,\-]+)', line)
            if vlan_match:
                vlan = vlan_match.group(1)
    return mode, vlan

def parse_ios_config(config_lines):
    """
    Parses Cisco IOS-XE configuration lines to extract key information.
    """
    parsed_data = {
        "hostname": None,
        "interfaces": [],
        "routing": [],
        "vlans": [],
        "ntp": [],
        "snmp": [],
        "syslog": [],
        "radius": [],
    }

    current_interface = None
    current_radius_server_name = None 

    i = 0
    while i < len(config_lines):
        original_line = config_lines[i] 
        line = original_line.strip() 

        # --- DEBUG PRINT: What line is being processed? ---
        # print(f"\nDEBUG: Processing line {i+1}: '{original_line.strip()}' (Stripped: '{line}')")
        # print(f"DEBUG: current_radius_server_name (before checks): {current_radius_server_name}")
        # print(f"DEBUG: current_interface (before checks): {current_interface['name'] if current_interface else 'None'}")
        # --- End DEBUG PRINT ---

        # --- Context Management: Clear contexts when new major blocks or separators are hit ---
        # This is CRITICAL for preventing lines from being mis-attributed.
        # If we hit a '!', it means the previous block has ended.
        # If we hit a top-level command that starts a *new* block, the old contexts should be cleared.
        is_major_block_start = any(line.startswith(cmd) for cmd in ["!", "hostname", "interface", "router", "vlan", "snmp-server host", "logging host", "ntp server", "radius server", "radius-server host"])

        if is_major_block_start:
            # If we were in an interface block and this new line is NOT a continuation of that interface,
            # then close out the current interface.
            # The specific 'interface' line itself will be handled by the 'elif line.startswith("interface")'
            if current_interface and not line.startswith("interface"):
                parsed_data["interfaces"].append(current_interface)
                current_interface = None # Clear interface context

            # Always clear RADIUS context if moving out of a specific radius server block
            # unless the current line is the start of a new radius server block.
            if current_radius_server_name is not None and not line.startswith("radius server"):
                current_radius_server_name = None
        # --- End Context Management ---
        
        if line.startswith("hostname"):
            parsed_data["hostname"] = line.split()[1]

        elif line.startswith("interface"):
            # Handle the start of a new interface block.
            # current_interface would already be cleared by the context management 'if' above
            # if this line is the start of a NEW interface block.
            current_interface = {"name": line.split()[1], "lines": []}

        elif current_interface: # This block now ONLY runs if we are genuinely inside an interface context
            current_interface["lines"].append(line)

        elif line.startswith("router"):
            parsed_data["routing"].extend(extract_ip_addresses([line]))

        elif line.startswith("vlan"):
            parsed_data["vlans"].append(line)

        elif line.startswith("ntp server"): 
            # print(f"DEBUG: >>> Matched 'ntp server' on line: '{line}'")
            extracted_ips = extract_ip_addresses([line])
            parsed_data["ntp"].extend(extracted_ips)
            # print(f"DEBUG: Extracted IPs for NTP: {extracted_ips}")
            # print(f"DEBUG: parsed_data['ntp'] after update: {parsed_data['ntp']}")

        elif line.startswith("snmp-server host"):
            parsed_data["snmp"].extend(extract_ip_addresses([line]))

        elif line.startswith("logging host"):
            parsed_data["syslog"].extend(extract_ip_addresses([line]))

        # --- RADIUS Parsing Logic ---
        elif line.startswith("radius server"):
            name_match = re.search(r'radius server (\S+)', line)
            if name_match:
                current_radius_server_name = name_match.group(1)
                # print(f"DEBUG: Matched 'radius server'. Set current_radius_server_name to: {current_radius_server_name}")
            else:
                current_radius_server_name = None 
        
        elif line.startswith("address ipv4") and current_radius_server_name: 
            ip_match = re.search(r'address ipv4 (\d+\.\d+\.\d+\.\d+)', line)
            if ip_match:
                ip_address = ip_match.group(1)
                parsed_data["radius"].append((current_radius_server_name, ip_address))
                # print(f"DEBUG: Matched 'address ipv4'. Appended: ({current_radius_server_name}, {ip_address})")
                current_radius_server_name = None 
        
        elif line.startswith("radius-server host"):
            ip_match = re.search(r'(\d+\.\d+\.\d+\.\d+)', line)
            name_match = re.search(r'radius-server host (\S+)', line) 
            if ip_match:
                server_name = name_match.group(1) if name_match else "Unknown_Old_Format"
                parsed_data["radius"].append((server_name, ip_match.group(1)))
            current_radius_server_name = None 

        i += 1
    
    # After the loop, if there's any remaining interface configuration, add it
    if current_interface:
        parsed_data["interfaces"].append(current_interface)

    # --- DEBUG PRINT: Final parsed data for NTP and RADIUS ---
    # print(f"\nDEBUG: Final parsed_data['ntp'] before export: {parsed_data['ntp']}")
    # print(f"DEBUG: Final parsed_data['radius'] before export: {parsed_data['radius']}")
    # --- End DEBUG PRINT ---

    return parsed_data


def export_to_excel(parsed_data, output_file):
    """
    Exports the parsed configuration data to an Excel spreadsheet.
    It can append data to an existing file or create a new one.
    """
    if os.path.exists(output_file) and os.path.getsize(output_file) > 0:
        try:
            wb = openpyxl.load_workbook(output_file)
        except openpyxl.utils.exceptions.InvalidFileException:
            print(f"Warning: '{output_file}' exists but is not a valid Excel file. Creating a new one.")
            wb = openpyxl.Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

    ws_title = "All Switch Info"
    if ws_title not in wb.sheetnames:
        ws = wb.create_sheet(title=ws_title)
        headers = ["Descriptor"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
    else:
        ws = wb[ws_title]

    col_index = ws.max_column + 1
    hostname_cell = ws.cell(row=1, column=col_index)
    hostname_cell.value = parsed_data["hostname"] or f"Switch {col_index-1}"
    hostname_cell.font = Font(bold=True) 

    row_num = 2 

    def write_row(descriptor, value):
        nonlocal row_num
        ws.cell(row=row_num, column=1).value = descriptor
        ws.cell(row=row_num, column=col_index).value = value
        row_num += 1

    write_row("Hostname", parsed_data["hostname"] or "Not Found")

    ws.cell(row=row_num, column=1).value = "--- INTERFACES ---"
    ws.cell(row=row_num, column=1).font = Font(bold=True)
    row_num += 1

    for iface in parsed_data["interfaces"]:
        ip_lines = extract_ip_addresses(iface['lines'])
        ip_info = ", ".join(ip_lines) if ip_lines else "No IP"
        mode, vlan = determine_interface_mode(iface['lines'])
        
        description_match = next((re.search(r'description (.+)', line) for line in iface['lines'] if re.match(r'description ', line)), None)
        description = description_match.group(1) if description_match else "No Description"

        write_row(f"Interface {iface['name']}", f"IP: {ip_info} | Mode: {mode} | VLAN: {vlan} | Desc: {description}")

    if parsed_data["routing"]:
        ws.cell(row=row_num, column=1).value = "--- ROUTING IPs ---"
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        for i, ip in enumerate(parsed_data["routing"], start=1):
            write_row(f"Routing IP #{i}", ip)

    if parsed_data["vlans"]:
        ws.cell(row=row_num, column=1).value = "--- VLANs ---"
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        for i, vlan_line in enumerate(parsed_data["vlans"], start=1):
            vlan_id_match = re.search(r'vlan (\d+)', vlan_line)
            vlan_name_match = re.search(r'name (\S+)', vlan_line)
            vlan_id = vlan_id_match.group(1) if vlan_id_match else "N/A"
            vlan_name = vlan_name_match.group(1) if vlan_name_match else "N/A"
            write_row(f"VLAN {vlan_id}", f"Name: {vlan_name}")

    if parsed_data["ntp"]:
        ws.cell(row=row_num, column=1).value = "--- NTP SERVERS ---"
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        for i, ip in enumerate(parsed_data["ntp"], start=1):
            write_row(f"NTP Server #{i}", ip)

    if parsed_data["snmp"]:
        ws.cell(row=row_num, column=1).value = "--- SNMP HOSTS ---"
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        for i, ip in enumerate(parsed_data["snmp"], start=1):
            write_row(f"SNMP Host #{i}", ip)

    if parsed_data["syslog"]:
        ws.cell(row=row_num, column=1).value = "--- SYSLOG SERVERS ---"
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        for i, ip in enumerate(parsed_data["syslog"], start=1):
            write_row(f"Syslog Server #{i}", ip)

    if parsed_data["radius"]:
        ws.cell(row=row_num, column=1).value = "--- RADIUS HOSTS ---"
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        for i, (name, ip) in enumerate(parsed_data["radius"], start=1):
            write_row(f"RADIUS Host #{i}", f"{name} ({ip})")
    
    wb.save(output_file)


def main():
    """
    Main function to handle command-line arguments, read config, parse, and export.
    """
    if len(sys.argv) != 2:
        print("Usage: python3 excel_export.py <ios_config_file.txt>")
        sys.exit(1)

    file_path = sys.argv[1]

    if not os.path.isfile(file_path):
        print(f"Error: File '{file_path}' not found.")
        sys.exit(1)

    try:
        with open(file_path, 'r') as f:
            config_lines = f.readlines()
    except Exception as e:
        print(f"Error reading file '{file_path}': {e}")
        sys.exit(1)

    parsed = parse_ios_config(config_lines)
    output_file = "switch_config_export.xlsx"
    
    export_to_excel(parsed, output_file)
    print(f"Configuration exported to {output_file}")


if __name__ == "__main__":
    main()
